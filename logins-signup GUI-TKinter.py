from tkinter import *
import datetime
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl import load_workbook
workbook = Workbook()
sheet = workbook.active
users_data = []
ud = {

}
wb = load_workbook('localdb.xlsx')
sheet = wb.active
xx = datetime.datetime.now()
z = 0
# Error ReRun


def signup():
    def signup_check():

        if len(ter_username.get())==11:
            if (ter_password1.get().find('@')>0) or ter_password1.get().find('&')>0 or ter_password1.get()[-1].isdigit() == True:
                if ter_password1.get() == ter_password2.get():
                    if (int(ter_dob.get()[0:2])<32) and (int(ter_dob.get()[4:5]) < 13) and (int(ter_dob.get()[6:10])< 2022):
                        if xx.year-int(ter_dob.get()[6:10]) >= 21:
                            ud = {
                                "users_name": ter_name.get(),
                                'username': ter_username.get(),
                                'password': ter_password2.get(),
                                'dob': ter_dob.get()
                            }

                            messagebox.showinfo('Signup complete', "You have successfully signed up")
                            sign_up_ter.destroy()

                            users_data.append(ud)
                            print(users_data)
                            sheet_len = len(sheet['A'])
                            print("A" + str(sheet_len+2), type("A" + str(sheet_len+2)))
                            sheet['A'+str(sheet_len+1)] = ter_name.get()
                            sheet['B'+str(sheet_len+1)] = ter_username.get()
                            sheet['C'+str(sheet_len+1)] = ter_password2.get()
                            sheet['D'+str(sheet_len+1)] = ter_dob.get()
                            wb.save('localdb.xlsx')
                            print(sheet['A'])
                        else:
                            messagebox.showerror('Age',"You must be 21 yrs or older")
                    else:
                        messagebox.showerror('DOB',"You have entered the Date of Birth in invalid format")
                else:
                    messagebox.showerror('Password',"Password does not match.")
            else:
                messagebox.showerror('Password',"Enter valid password containing '@', '&' or has a number at the end:")
        else:
            messagebox.showerror('Username',"Username(Phone Number) must be 11 numbers")

    sign_up_ter = Tk()
    sign_up_ter.title("Signup Terminal")
    sign_up_ter.geometry('500x400')
    # Name
    ter_name = StringVar()
    sign_up_name_label = Label(sign_up_ter, text='Enter your name below: ', font=60).place(x = 110)
    sign_up_name_input = Entry(sign_up_ter, textvariable= ter_name, bg='pink', width='50').place(x=110,y=20)
    # Username
    ter_username = StringVar()
    sign_up_username_label = Label(sign_up_ter, text='Enter your username (Phone Number) below: ', font=60).place(x=110, y=60)
    sign_up_username_input = Entry(sign_up_ter, textvariable=ter_username, bg='pink', width='50').place(x=110, y=80)
    # Password
    ter_password1 = StringVar()
    sign_up_password_label = Label(sign_up_ter, text='Enter a password containing "@", "&" or has a number\n at the '
                                                     'end below: ', font=60).place(x=110,y=120)
    sign_up_password_input = Entry(sign_up_ter, textvariable=ter_password1, bg='pink', width='50').place(x=110, y=160)
    # Repeat Password
    ter_password2 = StringVar()
    sign_up_password1_label = Label(sign_up_ter, text='Repeat password below: ', font=60).place(x=110, y=200)
    sign_up_password1_input = Entry(sign_up_ter, textvariable=ter_password2, bg='pink', width='50').place(x=110, y=220)
    # DOB
    ter_dob = StringVar()
    sign_up_dob_label = Label(sign_up_ter, text='Enter DOB in format DD/MM/YYYY no spaces:', font=60).place(x=110, y=260)
    sign_up_dob_input = Entry(sign_up_ter, textvariable=ter_dob, bg='pink', width='50').place(x=110, y=280)
    # Checks
    # Error from here
    sign_up_submit = Button(sign_up_ter, text='Submit',  fg='red', width=10, command=signup_check).place(x=150, y=320)

    sign_up_ter.mainloop()


def login():
    def login_check():
        xr = 1
        sheet_lent = len(sheet['B'])
        while xr <= len(sheet["B"]):
            if (login_ter_id.get() == sheet['B'+str(xr)].value) and (login_ter_password.get() == sheet['C'+str(xr)].value):
                messagebox.showinfo('Welcome', "Welcome " + sheet['A'+str(xr)].value)
                break
            else:
                if len(sheet["B"])-xr == 0:
                    messagebox.showinfo("Error", "You have not signed up with this contact number")
                    break
            xr += 1

    login_ter = Tk()
    login_ter.title("Sign Up Terminal")
    login_ter.geometry('500x400')
    # User ID
    login_ter_id = StringVar()
    login_id_label = Label(login_ter, text='Enter your username (Phone number) below: ', font=60).place(x=110)
    login_id_input = Entry(login_ter, textvariable=login_ter_id, bg='pink', width='50').place(x=110, y=20)
    # User Password
    login_ter_password = StringVar()
    login_password_label = Label(login_ter, text='Enter your password below: ', font=60).place(x=110,y=60)
    login_password_input = Entry(login_ter, textvariable=login_ter_password, bg='pink', width='50').place(x=110, y=80)
    # Checks
    login_submit = Button(login_ter, text="Submit", fg='red', width=10, command=login_check).place(x=150,y=120)


def ter_check():

        if ter_var.get() == '3':
            print("Thank You for Using the Application.")
            terminal.destroy()

        elif ter_var.get() == '1':
            terminal.destroy()
            signup()
        elif ter_var.get() == '2':
            terminal.destroy()
            login()
        else:
            print("Not Included")




while True:
    terminal = Tk()
    terminal.title('Login/Signup Terminal')
    terminal.geometry('600x400')
    ter_var = StringVar()
    terminal_info = Label(terminal, text="Enter 1 for Sign-up\n2 for Log-in\n3 for Quit.", font=60).place(x=110)
    terminal_123 = Entry(terminal, textvariable=ter_var, bg='pink', width='15').place(x=110, y=80)
    ter_submit = Button(terminal, text='Submit', fg='red', width=10, command=ter_check).place(x=110, y=110)
    if ter_var.get() == '3':
        break
    terminal.mainloop()