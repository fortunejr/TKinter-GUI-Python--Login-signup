from tkinter import *
import datetime
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl import load_workbook
import customtkinter


customtkinter.set_appearance_mode('dark')
customtkinter.set_default_color_theme('dark-blue')
sign_up_ter = customtkinter.CTk()
workbook = Workbook()
sheet = workbook.active
users_data = []
ud = {

}
wb = load_workbook('localdb.xlsx')
sheet = wb.active
xx = datetime.datetime.now()
z = 0


# Error ReRun


def signup():
    def signup_check():

        if len(ter_username.get()) == 11:
            if (ter_password1.get().find('@') > 0) or ter_password1.get().find('&') > 0 or ter_password1.get()[
                -1].isdigit() == True:
                if ter_password1.get() == ter_password2.get():
                    if (int(ter_dob.get()[0:2]) < 32) and (int(ter_dob.get()[4:5]) < 13) and (
                            int(ter_dob.get()[6:10]) < 2022):
                        if xx.year - int(ter_dob.get()[6:10]) >= 21:
                            ud = {
                                "users_name": ter_name.get(),
                                'username': ter_username.get(),
                                'password': ter_password2.get(),
                                'dob': ter_dob.get()
                            }

                            messagebox.showinfo('Signup complete', "You have successfully signed up")
                            sign_up_ter.destroy()

                            users_data.append(ud)
                            print(users_data)
                            sheet_len = len(sheet['A'])
                            print("A" + str(sheet_len + 2), type("A" + str(sheet_len + 2)))
                            sheet['A' + str(sheet_len + 1)] = ter_name.get()
                            sheet['B' + str(sheet_len + 1)] = ter_username.get()
                            sheet['C' + str(sheet_len + 1)] = ter_password2.get()
                            sheet['D' + str(sheet_len + 1)] = ter_dob.get()
                            wb.save('localdb.xlsx')
                            print(sheet['A'])
                        else:
                            messagebox.showerror('Age', "You must be 21 yrs or older")
                    else:
                        messagebox.showerror('DOB', "You have entered the Date of Birth in invalid format")
                else:
                    messagebox.showerror('Password', "Password does not match.")
            else:
                messagebox.showerror('Password', "Enter valid password containing '@', '&' or has a number at the end:")
        else:
            messagebox.showerror('Username', "Username(Phone Number) must be 11 numbers")

    
    frame = customtkinter.CTkFrame(master=sign_up_ter)
    frame.pack(pady=20,padx=60,fill='both',expand=True)
    sign_up_ter.title("Signup Terminal")
    sign_up_ter.geometry('500x600')
    # Name
    ter_name = StringVar()
    sign_up_name_label = customtkinter.CTkLabel(master=frame, text='Enter your name below: ').pack(pady=12,padx=10)
    sign_up_name_input = customtkinter.CTkEntry(master=frame, textvariable=ter_name).pack(pady=12,padx=10)
    # Username
    ter_username = StringVar()
    sign_up_username_label = customtkinter.CTkLabel(master=frame, text='Enter your username (Phone Number) below: ').pack(pady=12,padx=10)
    sign_up_username_input = customtkinter.CTkEntry(master=frame, textvariable=ter_username).pack(pady=12,padx=10)
    # Password
    ter_password1 = StringVar()
    sign_up_password_label = customtkinter.CTkLabel(master=frame, text='Enter a password containing "@", "&" or has a number\n at the '
                                                     'end below: ').pack(pady=12,padx=10)
    sign_up_password_input = customtkinter.CTkEntry(master=frame, textvariable=ter_password1).pack(pady=12,padx=10)
    # Repeat Password
    ter_password2 = StringVar()
    sign_up_password1_label = customtkinter.CTkLabel(master=frame, text='Repeat password below: ').pack(pady=12,padx=10)
    sign_up_password1_input = customtkinter.CTkEntry(master=frame, textvariable=ter_password2).pack(pady=12,padx=10)
    # DOB
    ter_dob = StringVar()
    sign_up_dob_label = customtkinter.CTkLabel(master=frame, text='Enter DOB in format DD/MM/YYYY no spaces:').pack(pady=12,padx=10)
    sign_up_dob_input = customtkinter.CTkEntry(master=frame, textvariable=ter_dob).pack(pady=12,padx=10)
    # Checks
    # Error from here
    sign_up_submit = customtkinter.CTkButton(master=frame, text='Submit', command=signup_check).pack(pady=12,padx=10)

    sign_up_ter.mainloop()


def login():
    def login_check():
        xr = 1
        sheet_lent = len(sheet['B'])
        while xr <= len(sheet["B"]):
            if (login_ter_id.get() == sheet['B' + str(xr)].value) and (
                    login_ter_password.get() == sheet['C' + str(xr)].value):
                messagebox.showinfo('Welcome', "Welcome " + sheet['A' + str(xr)].value)
                break
            else:
                if len(sheet["B"]) - xr == 0:
                    messagebox.showinfo("Error", "You have not signed up with this contact number")
                    break
            xr += 1

    login_ter = customtkinter.CTk()
    frame = customtkinter.CTkFrame(master=login_ter)
    frame.pack(pady=20,padx=60,fill='both',expand=True)
    login_ter.title("Sign Up Terminal")
    login_ter.geometry('500x400')
    # User ID
    login_ter_id = StringVar()
    login_id_label = customtkinter.CTkLabel(master=frame, text='Enter your username (Phone number) below: ').pack(pady=12,padx=10)
    login_id_input = customtkinter.CTkEntry(master=frame, textvariable=login_ter_id).pack(pady=12,padx=10)
    # User Password
    login_ter_password = StringVar()
    login_password_label = customtkinter.CTkLabel(master=frame, text='Enter your password below: ').pack(pady=12,padx=10)
    login_password_input = customtkinter.CTkEntry(master=frame, textvariable=login_ter_password).pack(pady=12,padx=10)
    # Checks
    login_submit = customtkinter.CTkButton(master=frame, text="Submit", command=login_check).pack(pady=12,padx=10)
    login_ter.mainloop()


def ter_check():
    print(ter_var.get())
    if ter_var.get() == 3:
        print("Thank You for Using the Application.")
        terminal.destroy()
    elif ter_var.get() == 1:
        terminal.destroy()
        signup()
    elif ter_var.get() == 2:
        terminal.destroy()
        login()
    else:
        print("Not Included")


while True:
    terminal = customtkinter.CTk()
    terminal.title('Login/Signup Terminal')
    terminal.geometry('600x400')
    frame = customtkinter.CTkFrame(master=terminal)
    frame.pack(pady=20,padx=60,fill='both',expand=True)
    ter_var = IntVar(frame)
    terminal_info = customtkinter.CTkLabel(master=frame, text="Enter 1 for Sign-up\n2 for Log-in\n3 for Quit.").pack(pady=12,padx=10)
    terminal_123 = customtkinter.CTkEntry(master=frame, textvariable=ter_var).pack(pady=12,padx=10)
    ter_submit = customtkinter.CTkButton(master=frame, text='Submit', command=ter_check).pack(pady=12,padx=10)
    terminal.mainloop()
    if ter_var.get() == 3:
        break
    
